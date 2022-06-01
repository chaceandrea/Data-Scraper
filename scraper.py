from openpyxl import load_workbook
import json
import pandas as pd
from simple_salesforce import Salesforce, SalesforceLogin
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import re
import datetime
from pyzipcode import ZipCodeDatabase
from timeit import default_timer as timer


def login():
    loginInfo = json.load(open('logn.json'))
    sfusername = loginInfo['sfusername']
    sfpassword = loginInfo['sfpassword']
    crmusername = loginInfo['crmusername']
    crmpassword = loginInfo['crmpassword']
    security_token = loginInfo['security_token']
    domain = 'login'
    session_id, instance = SalesforceLogin(username=sfusername, password=sfpassword, security_token=security_token,
                                           domain=domain)
    sf = Salesforce(instance=instance, session_id=session_id)

    wb = load_workbook(filename='FILENAME HERE' + '.xlsx')
    ws = wb["TAB NAME HERE"]
    driver = webdriver.Firefox()
    driver.get("WEBSITE URL")

    usernamefield = driver.find_element(By.XPATH, '//*[@id="user_username"]')
    usernamefield.send_keys(crmusername)

    passwordfield = driver.find_element(By.XPATH, '//*[@id="user_password"]')
    passwordfield.send_keys(crmpassword)
    loginbutton = driver.find_element(By.XPATH, '/html/body/div/div/div[1]/div[3]/div/form/div[4]/input')
    loginbutton.submit()

    time.sleep(7)

    return sf, ws, wb, driver


def set_pandas_settings():
    pd.set_option('display.max_columns', 100)
    pd.set_option('display.max_rows', 500)
    pd.set_option('display.min_rows', 500)
    pd.set_option('display.max_colwidth', 150)
    pd.set_option('display.width', 120)
    pd.set_option('expand_frame_repr', True)


def make_query(sf):
    insurancequery = "SELECT Account__r.Name, crm_url__c, Account__r.Gender__c, Account__r.Age__c, Facility__c, " \
                 "Discharge_Reason__c, Account__r.Alumni__r.Sobriety_Status__c, Account__r.Alumni__r.Days_Clean_Sober__c " \
                 "FROM Treatment_Episode__c WHERE " \
                 "AdmissionsDateNoTime__c > 2021-09-15 AND " \
                 "(Primary_Payer__c = 'INSURANCE COMPANY HERE' OR Primary_Payer__c = 'INSURANCE COMPANY HERE') " \
                 "AND Status__c = 'Discharged' " \
                 "ORDER BY AdmissionsDateNoTime__c"
    response = sf.query(insurancequery)
    firstrecords = response.get('records')
    nextrecordsurl = response.get('nextRecordsUrl')

    while not response.get('done'):
        response = sf.query_more(nextrecordsurl, identifier_is_url=True)
        firstrecords.extend(response.get('records'))
        nextrecordsurl = response.get('nextRecordsUrl')
    df_records = pd.DataFrame(firstrecords)
    accountnames = df_records['Account__r'].apply(pd.Series).drop(labels='attributes', axis=1)
    df_records.drop(labels=['crm_url__c', 'attributes'], axis=1)
    sobrietystatuses = accountnames['Alumni__r'].apply(pd.Series).drop(labels='attributes', axis=1)
    oldsalesforcedata = pd.concat([accountnames, df_records], axis=1).drop(labels=['attributes', 'Account__r'],
                                                                           axis=1)

    salesforcedata = pd.concat([oldsalesforcedata, sobrietystatuses], axis=1).drop(labels=['Alumni__r'],
                                                                                   axis=1)

    return salesforcedata


def get_days_authorized(driver, detoxdaysauthorized, rtcdaysauthorized, phpdaysauthorized, iopdaysauthorized):
    time.sleep(1)
    for x in driver.find_elements(By.XPATH, "//*[@class='page_break minheight160']"):
        loc = x.find_element(By.XPATH, ".//label[text()='Level of care']/parent::div").text
        loc = loc.split("\n")
        locdays = x.find_element(By.XPATH, ".//label[text()='# of days']/parent::div").text
        locdays = locdays.split("\n")
        status = x.find_element(By.XPATH, ".//label[text()='Status']/parent::div").text
        status = status.split("\n")
        insurance = x.find_element(By.XPATH, ".//label[text()='Insurance']/parent::div").text
        insurance = insurance.split("\n")
        if len(locdays) > 1 and len(loc) > 1 and len(status) > 1 and len(insurance) > 1:
            locdays = locdays[1]
            loc = loc[1]
            status = status[1]
            insurance = insurance[1]
            if 'Detox' in loc and status == "Approved" and ("INSURANCE COMPANY HERE" in insurance.upper()):
                detoxdaysauthorized += int(locdays)
            elif 'Residential' in loc and status == "Approved" and ("INSURANCE COMPANY HERE" in insurance.upper()):
                rtcdaysauthorized += int(locdays)
            elif 'PHP' in loc and status == "Approved" and ("INSURANCE COMPANY HERE" in insurance.upper()):
                phpdaysauthorized += int(locdays)
            elif 'IOP' in loc and status == "Approved" and ("INSURANCE COMPANY HERE" in insurance.upper()):
                iopdaysauthorized += int(locdays)
        else:
            continue
    return detoxdaysauthorized, rtcdaysauthorized, phpdaysauthorized, iopdaysauthorized


def get_discharge_reason(driver):
    dischargereason = driver.find_elements(By.XPATH, "//*[@class='ptop1em pright20px']")[1].text.split("\n")

    if len(dischargereason) > 1:
        dischargereason = dischargereason[1]
        if "Left voluntarily before completing treatment" in dischargereason:
            dischargereason = "Left voluntarily before completing treatment"
    else:
        dischargereason = ""
        dischargeerrorlist.append(row[1])
    return dischargereason


def get_state(driver):
    stateline = driver.find_element(By.XPATH, "//*[contains(text(), 'Current Address:')]").text.split("\n")[-1]
    zipcode = re.findall(r"(?<!\d)\d{5}(?!\d)", stateline)[0]
    zcdb = ZipCodeDatabase()
    state = zcdb[zipcode].state
    return state


def get_dx_codes(driver):
    try:
        dxbox = driver.find_element(By.XPATH, "//*[@class='patient_diagnosis_box']").text.split("\n")

        # Substance codes and dxcodes below
        whichdx = []
        for b in dxbox:
            for a in subcodes:
                if (a.lower() in b or a.upper() in b) and (a.upper() not in whichdx):
                    whichdx.append(a.upper())
        primarydx = whichdx[0]

        if len(whichdx) > 1:
            polysub = "Y"
            whichdx = whichdx[1:]
            whichdx = ', '.join(whichdx)
        else:
            polysub = "N"
            whichdx = ""

        # Co-Occuring Dx code below
        codx = []
        for b in dxbox:
            for a in cocodes:
                if (a.lower() in b or a.upper() in b) and (a.upper() not in codx):
                    codx.append(a.upper())
        if len(codx) > 1:
            codx = ', '.join(codx)
        elif len(codx) == 1:
            codx = codx[0]
        elif len(codx) == 0:
            codx = ""

        # Medical Dx code below
        meddx = []
        for b in dxbox:
            for a in medcodes:
                if (a.lower() in b or a.upper() in b) and (a.upper() not in meddx):
                    meddx.append(a.upper())
        if len(meddx) > 1:
            meddx = ', '.join(meddx)
        elif len(meddx) == 1:
            meddx = meddx[0]
        elif len(meddx) == 0:
            meddx = ""
    except:
        whichdx = ""
        meddx = ""
        codx = ""
        primarydx = ""
        polysub = "N"
        pass
    return whichdx, codx, meddx, primarydx, polysub


def get_er_past_12_months():
    try:
        erpast12months = driver.find_element(By.XPATH, "//*[contains(text(), 'In the past 12 months have you been "
                                                       "hospitalized or gone to the ER as a result of your chemical "
                                                       "dependency?')]/parent::div").text.split("\n")
        if len(erpast12months) > 1:
            erpast12months = erpast12months[1].upper()
            if erpast12months == 'NO':
                erpast12months = 'N'
            elif erpast12months == 'YES':
                erpast12months = 'Y'
            elif "IF YES, EXPLAIN" in erpast12months:
                erpast12months = erpast12months.split("IF YES, EXPLAIN:")[0]
                if 'NO' in erpast12months.upper():
                    erpast12months = "N"
                elif 'YES' in erpast12months.upper():
                    erpast12months = 'Y'
        else:
            erpast12months = ""
    except:
        erpast12months = ""
    return erpast12months


def get_education_level():
    try:
        educationlevel = driver.find_element(By.XPATH,
                                             "//*[contains(text(), 'Education:')]/parent::div").text.split("\n")
    except:
        educationlevel = driver.find_elements(By.XPATH, "//*[contains(text(), 'Educational')][1]/parent::div")[
            1].text.split("\n")

    if len(educationlevel) > 1:
        educationlevel = educationlevel[1].upper()
    else:
        educationlevel = ""
    return educationlevel


def get_longest_period_of_abstinence():
    longestpd = driver.find_element(By.XPATH, "//*[contains(text(), 'longest period of abstinence')]")
    longestpd = longestpd.find_element(By.XPATH, "..").text.split("\n")[1]
    return longestpd


start = timer()
nodaysauthorized = []
nodaysauthorizedcrm = []
prescreenerror = []
kipuurl = []
dischargeerrorlist = []
subcodes = ["F10", "F11", "F12", "F13", "F14", "F15", "F16", "F18", "F19"]
cocodes = ["F43", "F41", "F60.3", "F90", "F43.23", "F31", "F32", "F33", "F84", "F28", "F50", "F42"]
medcodes = ["R03.0", "I10", "E78", "E66", "K85", "K70", "G47", "J69", "B18", "E11", "R73.03", "K21", "L03.90", "I50.89",
            "R60", "B20", "J45", "M79.7", "C71.2", "E03", "E07", "N18", "K58", "K50", "G43", "G44", "G32", "L40.9",
            "L93.0", "G40", "F95.2", "G51", "Z33"]
cellrow = 1283
iteration = 0
formatdate = '%m/%d/%Y'

sf, ws, wb, driver = login()
set_pandas_settings()
salesforcedata = make_query(sf)

for count, row in enumerate(salesforcedata.itertuples()):
    detoxdaysauthorized = rtcdaysauthorized = phpdaysauthorized = iopdaysauthorized = 0
    preadmission = str(row[4] + "/records?process=11")
    clinicaltx = str(row[4] + "/records?process=7")

    driver.get(str(row[4]))
    detoxdaysauthorized, rtcdaysauthorized, phpdaysauthorized, iopdaysauthorized = \
        get_days_authorized(driver, detoxdaysauthorized, rtcdaysauthorized, phpdaysauthorized, iopdaysauthorized)
    if detoxdaysauthorized < 1 and rtcdaysauthorized < 1 and phpdaysauthorized < 1 and iopdaysauthorized < 1:
        continue

    admissiondate = driver.find_element(By.XPATH, "//*[@class='nowrap pright35px']/div[2]").text.split()[0]
    dischargedate = driver.find_element(By.XPATH, "//*[@class='nowrap ptop1em pright20px']/div[2]").text.split()[0]
    mrnumber = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[1]/div[2]/div/div[1]/div["
                                             "2]/table/tbody/tr/td[1]/h1/span[3]").text
    dischargereason = get_discharge_reason(driver)
    state = get_state(driver)
    los = driver.find_element(By.XPATH,
                              "/html/body/div[1]/div[3]/div[1]/div[2]/div[1]/div[1]/div[2]/table/tbody/tr/td[1]/p[4]/span").text

    whichdx, meddx, codx, primarydx, polysub = get_dx_codes(driver)

    try:
        kiputxepisodes = driver.find_element(By.XPATH, "//*[@name='case_files']").text.split("\n")
        kiputxepisodes = [x for x in kiputxepisodes if "YMed" not in (x)]
        if len(kiputxepisodes) > 1:
            for i in kiputxepisodes:
                if mrnumber in i:
                    indexnumber = kiputxepisodes.index(i)
            if (len(kiputxepisodes) - 1) > indexnumber:
                if (("AUS" not in kiputxepisodes[indexnumber + 1]) or ("ENC" not in mrnumber)):
                    firsttimehere = "N"
                    previousruadmits = str((len(kiputxepisodes) - (indexnumber + 1)))
                    lastadmit = kiputxepisodes[indexnumber + 1].split()[2]
                    admissiondatetime = datetime.datetime.strptime(admissiondate, formatdate)
                    lastadmitdatetime = datetime.datetime.strptime(lastadmit, formatdate)
                    dayssincelastadmit = str((admissiondatetime - lastadmitdatetime).days)
                else:
                    driver.find_element(By.XPATH, f"//*[contains(text(), '{kiputxepisodes[indexnumber + 1]}')]").click()
                    stepdowndate = driver.find_element(By.XPATH, "//*[@class='nowrap ptop1em pright20px']/div[2]").text
                    stepdowndate = stepdowndate.split()
                    stepdowndate = stepdowndate[0]
                    admissiondatetime = datetime.datetime.strptime(admissiondate, formatdate)
                    stepdowndatetime = datetime.datetime.strptime(stepdowndate, formatdate)
                    if (admissiondatetime - stepdowndatetime).days != 0:
                        firsttimehere = "N"
                        previousruadmits = str((len(kiputxepisodes) - (indexnumber + 1)))
                        lastadmit = kiputxepisodes[indexnumber + 1].split()[2]
                        admissiondatetime = datetime.datetime.strptime(admissiondate, formatdate)
                        lastadmitdatetime = datetime.datetime.strptime(lastadmit, formatdate)
                        dayssincelastadmit = str((admissiondatetime - lastadmitdatetime).days)
                    else:
                        if (len(kiputxepisodes) - 1) > indexnumber + 1:
                            firsttimehere = "N"
                            lastadmit = kiputxepisodes[indexnumber + 2].split()[2]
                            admissiondatetime = datetime.datetime.strptime(admissiondate, formatdate)
                            lastadmitdatetime = datetime.datetime.strptime(lastadmit, formatdate)
                            dayssincelastadmit = str((admissiondatetime - lastadmitdatetime).days)
                            stepdown = "STEPDOWN"
                            print(stepdown)
                        else:
                            firsttimehere = "Y"
                            dayssincelastadmit = ""

                    driver.get(str(row[4]))


            else:
                firsttimehere = "Y"
                dayssincelastadmit = ""
        else:
            firsttimehere = "Y"
            dayssincelastadmit = ""

    except:
        firsttimehere = "Y"
        dayssincelastadmit = ""
        previousruadmits = 0
        kiputxepisodes = []

    try:
        driver.get(preadmission)
        driver.get(driver.find_element(By.XPATH, "//*[contains(text(), 'Admission Screening')]").get_attribute('href'))
        # Now inside pre-screen

        educationlevel = get_education_level()

        erpast12months = get_er_past_12_months()

        longestpd = get_longest_period_of_abstinence()

        beenintxbefore = driver.find_element(By.XPATH,
                                             "//*[contains(text(), 'Has client been in treatment before?')]").find_element(
            By.XPATH,
            "..").text.split(
            "\n")
        if len(beenintxbefore) > 1:
            del beenintxbefore[0]
            beenintxbefore = beenintxbefore[0]
            if ", If yes, how many times?:" in beenintxbefore:
                howmanytimes = beenintxbefore.split(", If yes, how many times?:", 1)[1]
                beenintxbefore = beenintxbefore.split(", If yes, how many times?:", 1)[0]
                howmanytimes = re.findall(r'\d+', howmanytimes)
                if howmanytimes:
                    howmanytimes = howmanytimes[0]
                else:
                    howmanytimes = ""
                if ("Yes" or "yes") in beenintxbefore:
                    firsttimeintx = "N"
                elif ("No" or "no") in beenintxbefore:
                    if previousruadmits == 0:
                        firsttimeintx = "Y"
                        howmanytimes = ""
                    else:
                        firsttimeintx = "N"
                        howmanytimes = previousruadmits
            elif ("Yes" or "yes") in beenintxbefore:
                firsttimeintx = "N"
                howmanytimes = ""
            elif ("No" or "no") in beenintxbefore:
                if previousruadmits == 0:
                    firsttimeintx = "Y"
                    howmanytimes = ""
                else:
                    firsttimeintx = "N"
                    howmanytimes = previousruadmits
        elif (len(beenintxbefore) == 1 and "Has client been in treatment before?") in beenintxbefore:
            if previousruadmits == 0:
                firsttimeintx = "Y"
                howmanytimes = ""
            else:
                firsttimeintx = "N"
                howmanytimes = previousruadmits



    except:
        longestpd = "[Error]"
        beenintxbefore = "[Error]"
        firsttimeintx = "[Error]"
        howmanytimes = "[Error]"
        prescreenerror.append(row[1])
        kipuurl.append(row[4])
        pass

    # Go to biopsychosocial
    driver.get(clinicaltx)
    pcl_score = ""

    try:
        biopsychosocials = driver.find_elements(By.XPATH, "//*[contains(text(), 'Biopsychosocial')]")
        biopsycholinks = []
        for i in biopsychosocials:
            if "UPDATE" not in i.text.upper():
                biopsycholinks.append((i.get_attribute('href')))
        driver.get(biopsycholinks[0])
        pcl_score = driver.find_element(By.XPATH, "//*[contains(text(), 'Total Score:')]/parent::div").text.split("\n")[
            1].strip("()")
    except:
        if (len(kiputxepisodes) - 1) > indexnumber:
            if pcl_score == "":
                for index, mrnumber in enumerate(kiputxepisodes):
                    if index > indexnumber and pcl_score == "":
                        driver.find_element(By.XPATH, f"//*[contains(text(), '{mrnumber}')]").click()
                        clinicaltx = str(driver.current_url + "/records?process=7")
                        driver.get(clinicaltx)
                        biopsychosocials = driver.find_elements(By.XPATH, "//*[contains(text(), 'Biopsychosocial')]")
                        biopsycholinks = []
                        for i in biopsychosocials:
                            if "UPDATE" not in i.text.upper():
                                biopsycholinks.append((i.get_attribute('href')))
                        try:
                            driver.get(biopsycholinks[0])
                        except:
                            continue
                        try:
                            pcl_score = \
                                driver.find_element(By.XPATH,
                                                    "//*[contains(text(), 'Total Score:')]/parent::div").text.split(
                                    "\n")[1].strip("()")
                            if educationlevel == "":
                                educationlevel = \
                                    driver.find_element(By.XPATH, "//*[contains(text(), 'Educational')]/parent::div")[
                                        1].text.split(
                                        "\n")
                                if len(educationlevel) > 1:
                                    educationlevel = educationlevel[1].upper()
                        except:
                            continue
        pass

    # SALESFORCE LOGIC BELOW

    if str(row[2]) == "Male":
        gender = "M"
    else:
        gender = "F"

    # Format Data for Excel Sheet

    if ("MASTERS" in educationlevel.upper()) or ("DOCTORATE" in educationlevel.upper()) or ("POST COLLEGE DEGREE" in
                                                                                            educationlevel.upper()):
        educationlevel = "Post College Degree"
    elif ("COLLEGE DEGREE" in educationlevel.upper()) or ("ASSOCIATE" in educationlevel.upper() or ("BACHELOR" in
                                                                                                    educationlevel.upper())):
        educationlevel = "College Degree"
    elif "SOME COLLEGE" in educationlevel.upper():
        educationlevel = "Some College"
    elif (("HIGH SCHOOL" in educationlevel.upper()) or ("GED" in educationlevel.upper()) or ("G.E.D") in
          educationlevel.upper()) or ("12" in educationlevel.upper()) and (("IN HS" not in educationlevel.upper()) and
                                                                           ("IN HIGH" not in educationlevel.upper())):
        educationlevel = "High School"
    elif ("11" in educationlevel.upper()) or ("10" in educationlevel.upper()) or ("9" in educationlevel.upper()) or \
            ("IN HS" in educationlevel.upper()) or ("IN HIGH" in educationlevel.upper()):
        educationlevel = "Less than HS"

    if howmanytimes != "" and howmanytimes != "[Error]":
        howmanytimes = int(howmanytimes)

    if row[7] == 'Clean/Sober':
        sobstatus = 'Clean'
    elif row[7] == 'Deceased':
        sobstatus = 'Deceased'
    elif row[7] == 'Do Not Contact':
        sobstatus = 'UNK'
    elif row[7] == 'Functioning Addict':
        sobstatus = 'Clean'
    elif row[7] == 'Incarcerated':
        sobstatus = 'Incarcerated'
    elif row[7] == 'In Treatment elsewhere':
        sobstatus = 'In Tx - Other'
    elif row[7] == 'In Treatment here':
        sobstatus = 'In Tx - here'
    elif row[7] == 'MAT':
        sobstatus = 'Clean'
    elif row[7] == 'Unable to Contact':
        sobstatus = 'UNK'
    elif row[7] == 'Unknown':
        sobstatus = 'UNK'
    elif row[7] == 'Using':
        sobstatus = 'Using'

    if dayssincelastadmit != "":
        dayssincelastadmit = int(dayssincelastadmit)

    try:
        daysclean = int(row[8])
    except:
        daysclean = ""

    try:
        pcl_score = int(pcl_score)
    except:
        pcl_score = ""

    if detoxdaysauthorized > 0 or rtcdaysauthorized > 0 or phpdaysauthorized > 0 or iopdaysauthorized > 0:
        print(row[1] + "\n" + row[4])
        print("Gender:", gender)
        print("Age:", str(int(row[3])))
        print("Education Level:", educationlevel)
        print("State of Residence:", state)
        print("Facility:", str(row[5]))
        print("Admission Date: " + admissiondate)
        print("Length of stay: " + str(los.strip("()")))
        print("Detox Days Authorized: " + str(detoxdaysauthorized))
        print("Residential Days Authorized: " + str(rtcdaysauthorized))
        print("PHP Days Authorized: " + str(phpdaysauthorized))
        print("IOP Days Authorized: " + str(iopdaysauthorized))
        print("Primary Diagnosis: " + str(primarydx))
        print("Polysub: " + str(polysub))
        print("If yes, which dx?: " + str(whichdx))
        print("Co-Occurring: " + str(codx))
        print("Medical Dx: " + str(meddx))
        print("PCL Score:", str(pcl_score))
        print("ER past 12 Mo:", str(erpast12months))
        print("Longest Pd of Sobriety (Days): " + str(longestpd))
        print("Current Sobriety Status: " + sobstatus)
        print("If Clean, # of Days: " + str(daysclean))
        print("1st Time in Tx: " + str(firsttimeintx))
        print("If No, # of Past Tx: " + str(howmanytimes))
        print("First time here: " + str(firsttimehere))
        print("If No, Days Since Last Admit:", dayssincelastadmit)
        print("Discharge Date:", dischargedate)
        print("Discharge Reason:", dischargereason)
        print("\nIteration:", count, "out of", len(salesforcedata.index))
        print("Printed:", iteration)
        print("\n\n")

        # Update Excel Sheet

        ws.cell(row=cellrow, column=1).value = str(row[1])
        ws.cell(row=cellrow, column=2).value = str(gender)
        ws.cell(row=cellrow, column=3).value = int(row[3])
        ws.cell(row=cellrow, column=4).value = str(educationlevel)
        ws.cell(row=cellrow, column=5).value = str(state)
        ws.cell(row=cellrow, column=6).value = str(row[5])
        ws.cell(row=cellrow, column=7).value = str(admissiondate)
        ws.cell(row=cellrow, column=8).value = int(detoxdaysauthorized)
        ws.cell(row=cellrow, column=9).value = int(rtcdaysauthorized)
        ws.cell(row=cellrow, column=10).value = int(phpdaysauthorized)
        ws.cell(row=cellrow, column=11).value = int(iopdaysauthorized)
        ws.cell(row=cellrow, column=16).value = int(str(los.strip("()")))
        ws.cell(row=cellrow, column=17).value = str(primarydx)
        ws.cell(row=cellrow, column=18).value = str(polysub)
        ws.cell(row=cellrow, column=19).value = str(whichdx)
        ws.cell(row=cellrow, column=20).value = str(codx)
        ws.cell(row=cellrow, column=21).value = str(meddx)
        ws.cell(row=cellrow, column=22).value = pcl_score
        ws.cell(row=cellrow, column=23).value = erpast12months
        ws.cell(row=cellrow, column=28).value = str(longestpd)
        ws.cell(row=cellrow, column=29).value = sobstatus
        ws.cell(row=cellrow, column=30).value = daysclean
        ws.cell(row=cellrow, column=31).value = str(firsttimeintx)
        ws.cell(row=cellrow, column=32).value = howmanytimes
        ws.cell(row=cellrow, column=33).value = str(firsttimehere)
        ws.cell(row=cellrow, column=34).value = dayssincelastadmit
        ws.cell(row=cellrow, column=35).value = str(dischargedate)
        ws.cell(row=cellrow, column=36).value = str(dischargereason)

        iteration += 1
        cellrow += 1
        wb.save("FILENAME HERE" + '.xlsx')
    else:
        nodaysauthorized.append(row[1])
        nodaysauthorizedcrm.append(row[4])

print('program completed and data added for: ', iteration, 'clients')

if len(dischargeerrorlist) > 0:
    print("\n\n[Error] No discharge reason found for the following. Field left blank.")
    for i in (dischargeerrorlist):
        print(i)
if len(prescreenerror) > 0:
    print("\n\n[Error] Could not fetch pre-screen data for: ")
    for i, k in zip(prescreenerror, kipuurl):
        print(i)
        print(k)

elapsed_time = timer() - start
print("Program completed in:", (elapsed_time / 60), "minutes")
